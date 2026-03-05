import os
import tempfile
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook


def read_excel_sheets(path: str, sheets: Optional[list] = None) -> Dict[str, pd.DataFrame]:
    data = {}
    xls = pd.ExcelFile(path, engine="openpyxl")
    target_sheets = sheets or xls.sheet_names
    for name in target_sheets:
        if name in xls.sheet_names:
            data[name] = pd.read_excel(xls, sheet_name=name, engine="openpyxl")
    return data


def write_excel_preserve(input_path: str, output_path: str, sheets: Dict[str, pd.DataFrame]) -> None:
    # Atomic write to temp file in same directory
    out_dir = os.path.dirname(output_path)
    os.makedirs(out_dir or ".", exist_ok=True)
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=out_dir)
    os.close(fd)
    try:
        if os.path.exists(input_path):
            wb = load_workbook(input_path)
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                writer.book = wb
                for sheet_name, df in sheets.items():
                    if sheet_name in writer.book.sheetnames:
                        ws = writer.book[sheet_name]
                        writer.book.remove(ws)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
        else:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
        os.replace(temp_path, output_path)
    except PermissionError as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise PermissionError("Output file is locked or not writable") from e


